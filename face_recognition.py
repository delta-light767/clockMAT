import cv2
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime
import re

# Function to sanitize sheet titles
def sanitize_title(title):
    title = re.sub(r'[\\/*?[\]:]', ' ', title)  # Replace invalid characters
    return title[:31].strip()  # Limit to 31 characters

# Load the trained recognizer model and Haar Cascade
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read('trainer.yml')
face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')

# Load user data from CSV
user_data = pd.read_csv("user_data.csv")
user_dict = dict(zip(user_data["ID"], zip(user_data["Class"], user_data["Roll No"], user_data["Name"])))

# Initialize attendance Excel workbook
file_name = "attendance.xlsx"
try:
    workbook = load_workbook(file_name)
except FileNotFoundError:
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove default sheet

# Get current date for logging
current_date = datetime.now().strftime("%Y-%m-%d")

# Webcam initialization
cap = cv2.VideoCapture(0)
print("Press 'q' to quit and save attendance.")

while True:
    ret, frame = cap.read()
    if not ret:
        print("Error: Unable to access webcam.")
        break

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5)

    for (x, y, w, h) in faces:
        id, confidence = recognizer.predict(gray[y:y+h, x:x+w])

        # Check confidence threshold
        if confidence < 70:  # Adjust threshold as needed
            user_class, roll_no, name = user_dict.get(id, ("Unknown", "Unknown", "Unknown"))

            # Convert class to string and sanitize for valid sheet names
            sanitized_class = sanitize_title(str(user_class))

            # Create a new sheet for the class if it doesn't exist
            if sanitized_class not in workbook.sheetnames:
                workbook.create_sheet(title=sanitized_class)
                sheet = workbook[sanitized_class]

                # Add headers
                sheet.append(["Roll No", "Name", "Date", "Status"])
            else:
                sheet = workbook[sanitized_class]

            # Check if the student's attendance has already been marked
            rows = list(sheet.iter_rows(min_row=2, values_only=True))  # Exclude headers
            roll_numbers = [row[0] for row in rows]

            if roll_no in roll_numbers:
                # If already marked, display attendance taken message
                cv2.putText(frame, "Attendance Taken", (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
            else:
                # Add the student's attendance
                sheet.append([roll_no, name, current_date, "Present"])
                workbook.save(file_name)
                cv2.putText(frame, f"{name} Marked Present", (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2)

        else:
            # For unrecognized faces
            cv2.putText(frame, "Unknown", (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)

        # Draw rectangle around the face
        cv2.rectangle(frame, (x, y), (x+w, y+h), (255, 0, 0), 2)

    cv2.imshow("Face Recognition", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()
print("Attendance saved successfully.")
